"""Extract real Excel data into JSON seed files.

Reads:
  - C:/Users/User/Downloads/Supplier Price List .xlsx   (costing sheet: 1,468 SKUs w/ cost)
  - C:/Users/User/Downloads/Sales order details.xlsx    (1,654 SO line items)
  - C:/Users/User/Downloads/sales order.xlsx            (341 SO headers w/ address/phone/PO/remark)

Writes (to data-import/):
  - sku-master.json
  - so-lines.json
  - so-headers.json
"""
from __future__ import annotations
import json, sys, io, re, datetime, pathlib
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = pathlib.Path('C:/Users/User/Desktop/houzs-erp')
OUT = ROOT / 'data-import'
DL = pathlib.Path('C:/Users/User/Downloads')

# ─── Brand inference from item code prefix ────────────────────────────────────
BRAND_PREFIX = [
    ('AK-', 'AKEMI'), ('AK ', 'AKEMI'), ('AKE-', 'AKEMI'),
    ('ZNT-', 'ZANOTTI'), ('ZT-', 'ZANOTTI'), ('ZAN-', 'ZANOTTI'),
    ('DUN-', 'DUNLOPILLO'), ('DUN ', 'DUNLOPILLO'), ('DP-', 'DUNLOPILLO'),
    ('EGT-', 'ERGOTEX'), ('ERG-', 'ERGOTEX'), ('ERGO-', 'ERGOTEX'),
    ('HOK-', 'HOUZS'), ('HK-', 'HOUZS'),
    ('MYL-', 'MYLATEX'), ('ML-', 'MYLATEX'),
    ('GETHA-', 'GETHA'), ('GT-', 'GETHA'),
    ('NAKI', 'NAKI'),
    ('THL', 'THL3'),
    ('JM-', 'JM'), ('JMC-', 'JM'),
    ('AERO', 'AERO'),
    ('CARRESS', 'CARRESS'),
    ('NICOLLO', 'NICOLLO'),
    ('DORSETTLOFT', 'DORSETTLOFT'),
    ('ARMANI', 'ARMANI'),
    ('RED SOFA', 'RED_SOFA'),
    ('TNS-', 'TNS'),
    ('BEST-', 'BEST'),
    ('ANNEX', 'ANNEX'),
    ('ANNNEX', 'ANNEX'),
    ('MAJESTIC', 'MAJESTIC'),
    ('TODERN', 'TODERN'),
    ('LAVEO', 'LAVEO'),
    ('C&C', 'C_AND_C'),
    ('CC-', 'C_AND_C'),
]

def infer_brand(code: str) -> str:
    """Infer brand from SKU prefix. Returns 'OTHER' if no match."""
    if not code: return 'OTHER'
    c = code.strip().upper()
    for prefix, brand in BRAND_PREFIX:
        if c.startswith(prefix):
            return brand
    return 'OTHER'

def iso_date(v) -> str:
    if v is None: return ''
    if isinstance(v, datetime.datetime): return v.strftime('%Y-%m-%d')
    if isinstance(v, datetime.date): return v.strftime('%Y-%m-%d')
    if isinstance(v, str):
        try:
            return datetime.datetime.strptime(v.strip()[:10], '%Y-%m-%d').strftime('%Y-%m-%d')
        except: return str(v)[:10]
    return ''

def clean(v) -> str:
    if v is None: return ''
    s = str(v).strip()
    return s if s else ''

def num(v) -> float:
    if v is None or v == '': return 0.0
    try: return float(v)
    except: return 0.0

def nat(v) -> int:
    n = num(v)
    return max(1, int(n)) if n > 0 else 1

# ─── Extract SKU master from costing sheet ────────────────────────────────────
print('Reading Supplier Price List → costing sheet…')
wb = openpyxl.load_workbook(DL / 'Supplier Price List .xlsx', data_only=True)
ws = wb['costing']

skus = []
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
    code = clean(row[0])
    if not code: continue
    skus.append({
        'id': f'sku-{i:04d}',
        'itemCode': code,
        'uom': clean(row[1]) or 'UNIT',
        'supplier': clean(row[2]),
        'itemGroup': clean(row[3]) or 'OTHER',
        'description': clean(row[4]),
        'barCode': clean(row[5]),
        'costPrice': num(row[6]),
        'brand': infer_brand(code),
    })
print(f'  → {len(skus)} SKUs extracted')

# ─── Extract SO lines from Sales order details ────────────────────────────────
print('Reading Sales order details.xlsx…')
wb = openpyxl.load_workbook(DL / 'Sales order details.xlsx', data_only=True)
ws = wb['Sheet']

# SKU cost lookup
cost_by_code = { s['itemCode']: s['costPrice'] for s in skus }

so_lines = []
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
    doc_no = clean(row[1])
    if not doc_no: continue
    item_code = clean(row[17])
    qty = nat(row[22])
    unit_price = num(row[23])
    total = num(row[26])  # col 26 is 'Total' (pre-tax line total)
    cost = cost_by_code.get(item_code, 0.0) * qty
    so_lines.append({
        'id': f'sol-{i:05d}',
        'check': clean(row[0]),
        'docNo': doc_no,
        'date': iso_date(row[2]),
        'debtorCode': clean(row[3]),
        'debtorName': clean(row[4]),
        'agent': clean(row[5]),
        'currCode': clean(row[6]) or 'MYR',
        'itemGroup': clean(row[16]) or 'OTHERS',
        'itemCode': item_code,
        'description': clean(row[18]),
        'description2': clean(row[21]),
        'uom': clean(row[19]) or 'UNIT',
        'location': clean(row[20]),
        'qty': qty,
        'unitPrice': unit_price,
        'discount': num(row[24]),
        'total': total,
        'tax': num(row[27]),
        'totalExc': num(row[28]),
        'totalInc': num(row[29]),
        'creditorCode': clean(row[30]),
        'postToPO': clean(row[31]),
        'balance': num(row[32]),
        'payment': clean(row[33]) or 'Unchecked',
        'venue': clean(row[36]),
        'branding': clean(row[37]) or 'NONE',
        'cancelled': clean(row[12]) == 'T',
        'remark': clean(row[35]),
        'unitCost': cost_by_code.get(item_code, 0.0),
        'lineCost': cost,
        'lineMargin': total - cost,
    })
print(f'  → {len(so_lines)} SO lines extracted')

# ─── Extract SO headers from sales order.xlsx ────────────────────────────────
print('Reading sales order.xlsx…')
wb = openpyxl.load_workbook(DL / 'sales order.xlsx', data_only=True)
ws = wb['Sheet']

so_headers = []
for row in ws.iter_rows(min_row=2, values_only=True):
    doc_no = clean(row[0])
    if not doc_no: continue
    so_headers.append({
        'docNo': doc_no,
        'transferTo': clean(row[1]),
        'date': iso_date(row[2]),
        'branding': clean(row[3]) or 'NONE',
        'debtorName': clean(row[4]),
        'agent': clean(row[5]),
        'salesLocation': clean(row[6]),
        'ref': clean(row[7]),
        'localTotal': num(row[8]),
        'mattressSofa': num(row[9]),
        'bedframe': num(row[10]),
        'accessories': num(row[11]),
        'others': num(row[12]),
        'balance': num(row[13]),
        'remark2': clean(row[14]),
        'remark4': clean(row[15]),
        'remark3': clean(row[16]),
        'processingDate': iso_date(row[17]),
        'salesExemptionExpiry': iso_date(row[18]),
        'note': clean(row[19]),
        'poDocNo': clean(row[20]),
        'address1': clean(row[21]),
        'address2': clean(row[22]),
        'address3': clean(row[23]),
        'address4': clean(row[24]),
        'phone': clean(row[25]),
        'venue': clean(row[26]),
    })
print(f'  → {len(so_headers)} SO headers extracted')

# ─── Compute SO cost/margin roll-up ───────────────────────────────────────────
lines_by_doc: dict[str, list] = {}
for l in so_lines:
    lines_by_doc.setdefault(l['docNo'], []).append(l)

enriched_headers = []
for h in so_headers:
    ls = lines_by_doc.get(h['docNo'], [])
    total_cost = sum(l['lineCost'] for l in ls)
    total_rev = sum(l['total'] for l in ls)
    margin = total_rev - total_cost
    margin_pct = (margin / total_rev * 100) if total_rev > 0 else 0
    h['totalCost'] = round(total_cost, 2)
    h['totalRevenue'] = round(total_rev, 2)
    h['totalMargin'] = round(margin, 2)
    h['marginPct'] = round(margin_pct, 2)
    h['lineCount'] = len(ls)
    enriched_headers.append(h)

# ─── Write JSON seeds ─────────────────────────────────────────────────────────
OUT.mkdir(parents=True, exist_ok=True)
(OUT / 'sku-master.json').write_text(json.dumps(skus, ensure_ascii=False, indent=0), encoding='utf-8')
(OUT / 'so-lines.json').write_text(json.dumps(so_lines, ensure_ascii=False, indent=0), encoding='utf-8')
(OUT / 'so-headers.json').write_text(json.dumps(enriched_headers, ensure_ascii=False, indent=0), encoding='utf-8')

print(f'\nWrote:')
print(f'  {OUT / "sku-master.json"} ({len(skus)} SKUs)')
print(f'  {OUT / "so-lines.json"} ({len(so_lines)} lines)')
print(f'  {OUT / "so-headers.json"} ({len(enriched_headers)} headers)')

# ─── Summary stats ────────────────────────────────────────────────────────────
from collections import Counter
brands = Counter(s['brand'] for s in skus)
groups = Counter(s['itemGroup'] for s in skus)
total_revenue = sum(h['totalRevenue'] for h in enriched_headers)
total_cost = sum(h['totalCost'] for h in enriched_headers)
print(f'\nBrand inference:')
for b, n in brands.most_common():
    print(f'  {b}: {n}')
print(f'\nItem groups: {dict(groups)}')
print(f'\nGrand total revenue: RM {total_revenue:,.2f}')
print(f'Grand total cost: RM {total_cost:,.2f}')
print(f'Gross margin: RM {total_revenue - total_cost:,.2f} ({(total_revenue-total_cost)/total_revenue*100:.1f}%)')
