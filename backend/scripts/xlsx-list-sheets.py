import sys
from openpyxl import load_workbook
wb = load_workbook(sys.argv[1], data_only=True, read_only=True)
for s in wb.sheetnames:
    ws = wb[s]
    print(f"{s}\t{ws.max_row}x{ws.max_column}")
