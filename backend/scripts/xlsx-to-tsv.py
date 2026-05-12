#!/usr/bin/env python3
"""One-shot XLSX → TSV converter for the project seed flow.

Usage: python xlsx-to-tsv.py <xlsx-path> <tsv-output-path>

Reads the first sheet, formats each cell as plain text (numbers
stripped of trailing .0, dates kept verbatim), tab-separates, and
writes UTF-8 TSV. Conservatively quotes nothing — the seed parser
splits on \\t only.
"""
import sys
from openpyxl import load_workbook

if len(sys.argv) < 3:
    print("Usage: xlsx-to-tsv.py <xlsx> <tsv> [sheet-name]", file=sys.stderr)
    sys.exit(2)

src, dst = sys.argv[1], sys.argv[2]
sheet = sys.argv[3] if len(sys.argv) >= 4 else None
wb = load_workbook(src, data_only=True)
ws = wb[sheet] if sheet else wb.active

def fmt(v):
    if v is None:
        return ""
    if isinstance(v, float):
        # Drop trailing .0 on whole-number floats.
        return str(int(v)) if v.is_integer() else f"{v:.6f}".rstrip("0").rstrip(".")
    return str(v).replace("\t", " ").replace("\r", " ").replace("\n", " ")

with open(dst, "w", encoding="utf-8", newline="") as f:
    for row in ws.iter_rows(values_only=True):
        f.write("\t".join(fmt(c) for c in row) + "\n")

print(f"Wrote {dst}")
